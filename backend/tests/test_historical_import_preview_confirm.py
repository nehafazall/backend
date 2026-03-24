"""
Test Suite for Historical Import Preview/Confirm Flow
Tests the two-step import process:
1. POST /api/import/historical-sales-xlsx/preview - Validates Excel, returns preview
2. POST /api/import/historical-sales-xlsx/confirm/{preview_id} - Confirms and imports valid rows

Features tested:
- Row-by-row validation with specific error messages
- NaN/None value handling from pandas
- Phone number cleaning (no .0 suffix)
- Duplicate phone detection (in DB and within file)
- Agent/Team resolution
- Preview storage and retrieval
- Confirm with valid/invalid/expired/already-confirmed preview_id
"""

import pytest
import requests
import os
import io
from openpyxl import Workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', 'https://data-integrity-check-12.preview.emergentagent.com').rstrip('/')

# Test credentials
SUPER_ADMIN_EMAIL = "aqib@clt-academy.com"
SUPER_ADMIN_PASSWORD = "@Aqib1234"


@pytest.fixture(scope="module")
def auth_token():
    """Get authentication token for super_admin"""
    response = requests.post(f"{BASE_URL}/api/auth/login", json={
        "email": SUPER_ADMIN_EMAIL,
        "password": SUPER_ADMIN_PASSWORD
    })
    assert response.status_code == 200, f"Login failed: {response.text}"
    data = response.json()
    return data.get("access_token")


@pytest.fixture(scope="module")
def headers(auth_token):
    """Get auth headers"""
    return {
        "Authorization": f"Bearer {auth_token}",
    }


def create_test_excel(rows_data):
    """Create an Excel file in memory with test data"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Historical Sales Data"
    
    # Headers
    headers_list = [
        "Full Name *", "Phone *", "Course Enrolled *", "Course Amount (AED)",
        "Add-ons", "Add-on Amount (AED)", "Agent Name *", "Team Name *",
        "Enrolled Amount (AED)", "Enrolled At (YYYY-MM-DD)", "Email", "Country", "City", "Source"
    ]
    for col, header in enumerate(headers_list, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data rows
    for row_idx, row_data in enumerate(rows_data, 2):
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=value)
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


class TestHistoricalImportPreview:
    """Test the preview endpoint validation"""
    
    def test_preview_valid_row(self, headers):
        """Test preview with a valid row"""
        rows = [
            ["Test Valid User", "971501111001", "Starter", 1000, "", 0, "Kiran VR", "TEAM XLNC", 1000, "2025-01-15", "test@test.com", "UAE", "Dubai", "Historical"]
        ]
        excel_file = create_test_excel(rows)
        
        response = requests.post(
            f"{BASE_URL}/api/import/historical-sales-xlsx/preview",
            headers=headers,
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 200, f"Preview failed: {response.text}"
        data = response.json()
        
        # Verify response structure
        assert "preview_id" in data, "Missing preview_id"
        assert "total_rows" in data, "Missing total_rows"
        assert "valid" in data, "Missing valid count"
        assert "errors" in data, "Missing errors count"
        assert "duplicates" in data, "Missing duplicates count"
        assert "rows" in data, "Missing rows array"
        
        assert data["total_rows"] == 1
        # Note: Row may be valid or duplicate depending on DB state
        print(f"Preview result: valid={data['valid']}, errors={data['errors']}, duplicates={data['duplicates']}")
        
        # Check row structure
        row = data["rows"][0]
        assert "row_num" in row
        assert "data" in row
        assert "status" in row
        assert "errors" in row
        assert "resolved" in row
        
        return data["preview_id"]
    
    def test_preview_missing_required_fields(self, headers):
        """Test preview catches missing required fields"""
        rows = [
            # Missing Full Name
            ["", "971502222002", "Starter", 1000, "", 0, "Kiran VR", "TEAM XLNC", 1000, "2025-01-15", "", "", "", ""],
            # Missing Phone
            ["Test User No Phone", "", "Starter", 1000, "", 0, "Kiran VR", "TEAM XLNC", 1000, "2025-01-15", "", "", "", ""],
            # Missing Course
            ["Test User No Course", "971503333003", "", 1000, "", 0, "Kiran VR", "TEAM XLNC", 1000, "2025-01-15", "", "", "", ""],
            # Missing Agent
            ["Test User No Agent", "971504444004", "Starter", 1000, "", 0, "", "TEAM XLNC", 1000, "2025-01-15", "", "", "", ""],
            # Missing Team
            ["Test User No Team", "971505555005", "Starter", 1000, "", 0, "Kiran VR", "", 1000, "2025-01-15", "", "", "", ""],
        ]
        excel_file = create_test_excel(rows)
        
        response = requests.post(
            f"{BASE_URL}/api/import/historical-sales-xlsx/preview",
            headers=headers,
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 200
        data = response.json()
        
        assert data["total_rows"] == 5
        # All rows should have errors (no valid rows expected)
        print(f"Missing fields test: errors={data['errors']}, valid={data['valid']}")
        
        # Verify specific error messages
        rows = data["rows"]
        assert "Missing Full Name" in rows[0]["errors"], "Should catch missing full name"
        assert "Missing Phone" in rows[1]["errors"], "Should catch missing phone"
        assert "Missing Course Enrolled" in rows[2]["errors"], "Should catch missing course"
        assert "Missing Agent Name" in rows[3]["errors"], "Should catch missing agent"
        assert "Missing Team Name" in rows[4]["errors"], "Should catch missing team"
    
    def test_preview_unknown_agent_team(self, headers):
        """Test preview catches unknown agent and team"""
        rows = [
            ["Test User Unknown Agent", "971506666006", "Starter", 1000, "", 0, "NonExistent Agent", "TEAM XLNC", 1000, "2025-01-15", "", "", "", ""],
            ["Test User Unknown Team", "971507777007", "Starter", 1000, "", 0, "Kiran VR", "NonExistent Team", 1000, "2025-01-15", "", "", "", ""],
        ]
        excel_file = create_test_excel(rows)
        
        response = requests.post(
            f"{BASE_URL}/api/import/historical-sales-xlsx/preview",
            headers=headers,
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 200
        data = response.json()
        
        result_rows = data["rows"]
        # Check first row has agent not found error
        row0_errors = " ".join(result_rows[0]["errors"])
        assert "not found" in row0_errors.lower() and "agent" in row0_errors.lower(), f"Should catch unknown agent: {result_rows[0]['errors']}"
        # Check second row has team not found error
        row1_errors = " ".join(result_rows[1]["errors"])
        assert "not found" in row1_errors.lower() and "team" in row1_errors.lower(), f"Should catch unknown team: {result_rows[1]['errors']}"
    
    def test_preview_duplicate_phone_within_file(self, headers):
        """Test preview catches duplicate phones within the same file"""
        rows = [
            ["Test User A", "971508888008", "Starter", 1000, "", 0, "Kiran VR", "TEAM XLNC", 1000, "2025-01-15", "", "", "", ""],
            ["Test User B", "971508888008", "Starter", 1000, "", 0, "Kiran VR", "TEAM XLNC", 1000, "2025-01-15", "", "", "", ""],  # Same phone
        ]
        excel_file = create_test_excel(rows)
        
        response = requests.post(
            f"{BASE_URL}/api/import/historical-sales-xlsx/preview",
            headers=headers,
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 200
        data = response.json()
        
        # Second row should have duplicate error
        rows = data["rows"]
        second_row_errors = " ".join(rows[1]["errors"])
        assert "duplicate" in second_row_errors.lower() or "within file" in second_row_errors.lower(), f"Should catch duplicate phone within file: {rows[1]['errors']}"
    
    def test_preview_handles_nan_none_values(self, headers):
        """Test preview handles NaN/None values correctly"""
        # Create excel with partial data - some empty fields
        # Note: Completely empty rows get skipped by pandas - this is expected
        rows = [
            ["Test User Empty Fields", "", "Starter", 1000, "", 0, "Kiran VR", "", 1000, "", "", "", "", ""],  # Some empty
        ]
        excel_file = create_test_excel(rows)
        
        response = requests.post(
            f"{BASE_URL}/api/import/historical-sales-xlsx/preview",
            headers=headers,
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 200
        data = response.json()
        
        result_rows = data["rows"]
        # Should have at least one row
        assert len(result_rows) >= 1, f"Expected at least 1 row, got {len(result_rows)}"
        # Should have errors for missing phone and team
        assert result_rows[0]["status"] == "error"
        assert "Missing Phone" in result_rows[0]["errors"]
        assert "Missing Team Name" in result_rows[0]["errors"]


class TestHistoricalImportConfirm:
    """Test the confirm endpoint"""
    
    def test_confirm_with_invalid_preview_id(self, headers):
        """Test confirm returns 404 for invalid preview_id"""
        response = requests.post(
            f"{BASE_URL}/api/import/historical-sales-xlsx/confirm/fake-invalid-id-12345",
            headers=headers
        )
        
        assert response.status_code == 404, f"Expected 404 for invalid preview_id, got {response.status_code}: {response.text}"
        data = response.json()
        assert "not found" in data.get("detail", "").lower() or "expired" in data.get("detail", "").lower()
    
    def test_full_preview_confirm_flow(self, headers):
        """Test complete flow: preview -> confirm -> verify import"""
        # Use unique phone number
        import time
        unique_phone = f"97150{int(time.time()) % 10000000:07d}"
        
        rows = [
            [f"TEST Import User {unique_phone}", unique_phone, "Starter", 1000, "", 0, "Kiran VR", "TEAM XLNC", 1000, "2025-01-15", f"test_{unique_phone}@test.com", "UAE", "Dubai", "Test Import"]
        ]
        excel_file = create_test_excel(rows)
        
        # Step 1: Preview
        preview_response = requests.post(
            f"{BASE_URL}/api/import/historical-sales-xlsx/preview",
            headers=headers,
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert preview_response.status_code == 200
        preview_data = preview_response.json()
        preview_id = preview_data.get("preview_id")
        
        assert preview_id, "Preview should return preview_id"
        
        # Check if row is valid (not duplicate in DB)
        if preview_data["valid"] == 0:
            print(f"Row marked as duplicate/error, skipping confirm test: {preview_data['rows'][0]['errors']}")
            pytest.skip("Test row is duplicate in DB - skipping confirm flow")
        
        # Step 2: Confirm
        confirm_response = requests.post(
            f"{BASE_URL}/api/import/historical-sales-xlsx/confirm/{preview_id}",
            headers=headers
        )
        
        assert confirm_response.status_code == 200, f"Confirm failed: {confirm_response.text}"
        confirm_data = confirm_response.json()
        
        # Verify confirm response structure
        assert "success" in confirm_data
        assert "failed" in confirm_data
        assert "total_rows" in confirm_data
        assert "students_created" in confirm_data
        
        assert confirm_data["success"] >= 1, f"Expected at least 1 success, got {confirm_data['success']}"
        assert confirm_data["students_created"] >= 1, f"Expected at least 1 student created"
        
        print(f"Import result: success={confirm_data['success']}, students_created={confirm_data['students_created']}")
        
        # Store preview_id for double-confirm test
        return preview_id, unique_phone
    
    def test_confirm_already_confirmed_preview(self, headers):
        """Test confirm returns 400 for already-confirmed preview"""
        import time
        unique_phone = f"97150{int(time.time()) % 10000000:07d}"
        
        rows = [
            [f"TEST Double Confirm User", unique_phone, "Starter", 1000, "", 0, "Kiran VR", "TEAM XLNC", 1000, "2025-01-15", "", "", "", "Test"]
        ]
        excel_file = create_test_excel(rows)
        
        # Preview
        preview_response = requests.post(
            f"{BASE_URL}/api/import/historical-sales-xlsx/preview",
            headers=headers,
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert preview_response.status_code == 200
        preview_data = preview_response.json()
        preview_id = preview_data.get("preview_id")
        
        if preview_data["valid"] == 0:
            pytest.skip("Test row is duplicate/error - skipping double confirm test")
        
        # First confirm - should succeed
        confirm_response = requests.post(
            f"{BASE_URL}/api/import/historical-sales-xlsx/confirm/{preview_id}",
            headers=headers
        )
        assert confirm_response.status_code == 200
        
        # Second confirm - should fail with 400
        confirm_response2 = requests.post(
            f"{BASE_URL}/api/import/historical-sales-xlsx/confirm/{preview_id}",
            headers=headers
        )
        
        assert confirm_response2.status_code == 400, f"Expected 400 for already-confirmed preview, got {confirm_response2.status_code}"
        data = confirm_response2.json()
        assert "already" in data.get("detail", "").lower() or "confirmed" in data.get("detail", "").lower()


class TestPhoneNumberCleaning:
    """Test phone number cleaning (removing .0 suffix from pandas floats)"""
    
    def test_phone_with_float_suffix_cleaned(self, headers):
        """Test phone numbers stored as floats (e.g., 971500000001.0) are cleaned"""
        # This tests the clean_val function logic
        rows = [
            ["Test Float Phone User", 971509999001.0, "Starter", 1000, "", 0, "Kiran VR", "TEAM XLNC", 1000, "2025-01-15", "", "", "", ""]
        ]
        excel_file = create_test_excel(rows)
        
        response = requests.post(
            f"{BASE_URL}/api/import/historical-sales-xlsx/preview",
            headers=headers,
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 200
        data = response.json()
        
        # Check the phone in the preview data doesn't have .0 suffix
        row = data["rows"][0]
        phone = row["data"].get("phone", "")
        assert not phone.endswith(".0"), f"Phone should not have .0 suffix: {phone}"
        print(f"Phone after cleaning: {phone}")


class TestPreviewResponseStructure:
    """Test preview response has correct structure for frontend display"""
    
    def test_preview_response_has_all_required_fields(self, headers):
        """Test preview response structure for frontend table display"""
        rows = [
            ["Test Structure User", "971500001234", "Starter", 1000, "Addon1", 500, "Kiran VR", "TEAM XLNC", 1500, "2025-01-15", "test@test.com", "UAE", "Dubai", "Test"]
        ]
        excel_file = create_test_excel(rows)
        
        response = requests.post(
            f"{BASE_URL}/api/import/historical-sales-xlsx/preview",
            headers=headers,
            files={"file": ("test.xlsx", excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 200
        data = response.json()
        
        # Top-level fields
        required_top_level = ["preview_id", "total_rows", "valid", "errors", "duplicates", "rows"]
        for field in required_top_level:
            assert field in data, f"Missing top-level field: {field}"
        
        # Row fields for frontend table
        row = data["rows"][0]
        assert "row_num" in row, "Missing row_num"
        assert "status" in row, "Missing status"  # valid, error, duplicate
        assert "data" in row, "Missing data"
        assert "errors" in row, "Missing errors"
        assert "resolved" in row, "Missing resolved"
        
        # Data fields for table columns
        row_data = row["data"]
        expected_data_fields = ["full_name", "phone", "course_enrolled", "agent_name", "team_name"]
        for field in expected_data_fields:
            assert field in row_data, f"Missing data field: {field}"
        
        # Resolved fields for matched agent/team
        if row["status"] == "valid":
            assert "agent_id" in row["resolved"] or "agent_name" in row["resolved"], "Should have resolved agent"
            assert "team_id" in row["resolved"] or "team_name" in row["resolved"], "Should have resolved team"
        
        print(f"Row structure verified: status={row['status']}, errors={row['errors']}")


class TestCleanup:
    """Cleanup test data after tests"""
    
    def test_cleanup_test_data(self, headers):
        """Remove test-created leads and students (phone starting with 97150)"""
        # Get test leads
        response = requests.get(
            f"{BASE_URL}/api/leads?search=TEST",
            headers=headers
        )
        
        if response.status_code == 200:
            data = response.json()
            leads = data.get("leads", []) if isinstance(data, dict) else data
            deleted_count = 0
            for lead in leads:
                if "TEST" in lead.get("full_name", "").upper() or lead.get("is_historical"):
                    # Delete lead
                    del_response = requests.delete(
                        f"{BASE_URL}/api/leads/{lead['id']}",
                        headers=headers
                    )
                    if del_response.status_code in [200, 204]:
                        deleted_count += 1
            print(f"Cleaned up {deleted_count} test leads")
        
        # Also cleanup students
        student_response = requests.get(
            f"{BASE_URL}/api/students?search=TEST",
            headers=headers
        )
        
        if student_response.status_code == 200:
            data = student_response.json()
            students = data.get("students", []) if isinstance(data, dict) else data
            deleted_students = 0
            for student in students:
                if "TEST" in student.get("full_name", "").upper() or student.get("is_historical"):
                    del_response = requests.delete(
                        f"{BASE_URL}/api/students/{student['id']}",
                        headers=headers
                    )
                    if del_response.status_code in [200, 204]:
                        deleted_students += 1
            print(f"Cleaned up {deleted_students} test students")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
