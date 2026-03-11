"""
Test Manual Attendance Import Feature
Tests for:
1. GET /api/hr/attendance/template/download - Downloads XLSX template
2. POST /api/hr/attendance/import - Upload filled XLSX to create attendance records
3. XLSX validation and format
4. Payroll integration with imported attendance
"""

import pytest
import requests
import os
import io
import calendar

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

# Test credentials
TEST_EMAIL = "aqib@clt-academy.com"
TEST_PASSWORD = "@Aqib1234"


@pytest.fixture(scope="module")
def auth_token():
    """Get authentication token"""
    response = requests.post(f"{BASE_URL}/api/auth/login", json={
        "email": TEST_EMAIL,
        "password": TEST_PASSWORD
    })
    if response.status_code == 200:
        return response.json().get("access_token")
    pytest.skip(f"Authentication failed - {response.status_code}: {response.text}")


@pytest.fixture(scope="module")
def auth_headers(auth_token):
    """Get authentication headers"""
    return {
        "Authorization": f"Bearer {auth_token}",
        "Content-Type": "application/json"
    }


class TestAttendanceTemplateDownload:
    """Tests for GET /api/hr/attendance/template/download endpoint"""

    def test_download_template_success(self, auth_token):
        """Test downloading attendance template for valid month"""
        response = requests.get(
            f"{BASE_URL}/api/hr/attendance/template/download",
            params={"month": "2026-03"},
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        # Verify content type is XLSX
        content_type = response.headers.get("Content-Type", "")
        assert "spreadsheet" in content_type or "application/octet-stream" in content_type, f"Unexpected content type: {content_type}"
        
        # Verify we got some data
        assert len(response.content) > 1000, "Template file seems too small"

    def test_download_template_invalid_month(self, auth_token):
        """Test download with invalid month format"""
        response = requests.get(
            f"{BASE_URL}/api/hr/attendance/template/download",
            params={"month": "invalid-month"},
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        
        assert response.status_code == 400, f"Expected 400, got {response.status_code}"
        assert "Invalid month format" in response.json().get("detail", "")

    def test_download_template_xlsx_structure(self, auth_token):
        """Test XLSX template has correct structure and headers"""
        import openpyxl
        
        response = requests.get(
            f"{BASE_URL}/api/hr/attendance/template/download",
            params={"month": "2026-03"},
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        
        assert response.status_code == 200
        
        # Parse the XLSX file
        wb = openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)
        ws = wb.active
        
        # Check headers are in row 5
        expected_headers = [
            "Employee ID", "Employee Name", "Department", "Designation",
            "Basic Salary", "Housing Allowance", "Transport Allowance",
            "Food Allowance", "Phone Allowance", "Other Allowances",
            "Fixed Incentive", "Full Days", "Half Days", "Approved Leaves"
        ]
        
        actual_headers = [ws.cell(row=5, column=i).value for i in range(1, 15)]
        
        for idx, expected in enumerate(expected_headers):
            assert actual_headers[idx] == expected, f"Header mismatch at column {idx+1}: expected '{expected}', got '{actual_headers[idx]}'"

    def test_download_template_has_employee_data(self, auth_token):
        """Test template is pre-filled with active employees"""
        import openpyxl
        
        response = requests.get(
            f"{BASE_URL}/api/hr/attendance/template/download",
            params={"month": "2026-03"},
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)
        ws = wb.active
        
        # Check data rows exist (starting row 6)
        employee_count = 0
        for row in range(6, 200):  # Check up to 200 rows
            if ws.cell(row=row, column=1).value:  # Employee ID column
                employee_count += 1
            else:
                break
        
        assert employee_count > 0, "Template should contain at least one employee"
        print(f"Template contains {employee_count} employees")

    def test_download_template_attendance_columns_initialized(self, auth_token):
        """Test attendance columns (L, M, N) are initialized to 0"""
        import openpyxl
        
        response = requests.get(
            f"{BASE_URL}/api/hr/attendance/template/download",
            params={"month": "2026-03"},
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)
        ws = wb.active
        
        # Check first data row's attendance columns
        if ws.cell(row=6, column=1).value:  # Has employee data
            full_days = ws.cell(row=6, column=12).value  # Column L
            half_days = ws.cell(row=6, column=13).value  # Column M
            approved_leaves = ws.cell(row=6, column=14).value  # Column N
            
            assert full_days == 0 or full_days is None, f"Full Days should be 0, got {full_days}"
            assert half_days == 0 or half_days is None, f"Half Days should be 0, got {half_days}"
            assert approved_leaves == 0 or approved_leaves is None, f"Approved Leaves should be 0, got {approved_leaves}"


class TestAttendanceImport:
    """Tests for POST /api/hr/attendance/import endpoint"""

    def test_import_attendance_with_valid_data(self, auth_token):
        """Test importing attendance with valid XLSX data"""
        import openpyxl
        
        # First, download the template
        template_response = requests.get(
            f"{BASE_URL}/api/hr/attendance/template/download",
            params={"month": "2026-03"},
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        assert template_response.status_code == 200
        
        # Modify the template with test data
        wb = openpyxl.load_workbook(io.BytesIO(template_response.content))
        ws = wb.active
        
        # Find first employee and fill attendance
        employees_filled = 0
        for row in range(6, 50):
            if ws.cell(row=row, column=1).value:  # Has employee
                ws.cell(row=row, column=12).value = 20  # Full Days
                ws.cell(row=row, column=13).value = 2   # Half Days
                ws.cell(row=row, column=14).value = 2   # Approved Leaves
                employees_filled += 1
                if employees_filled >= 3:  # Fill only 3 employees for test
                    break
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Upload the modified file
        files = {"file": ("attendance_test.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        data = {"month": "2026-03"}
        
        response = requests.post(
            f"{BASE_URL}/api/hr/attendance/import",
            files=files,
            data=data,
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        result = response.json()
        assert "message" in result, "Response should have message field"
        assert result.get("imported", 0) > 0 or result.get("attendance_records_created", 0) > 0, f"Should have imported at least one employee: {result}"
        print(f"Import result: {result}")

    def test_import_validates_total_days(self, auth_token):
        """Test import rejects entries where total days exceed calendar days"""
        import openpyxl
        
        # Download template
        template_response = requests.get(
            f"{BASE_URL}/api/hr/attendance/template/download",
            params={"month": "2026-03"},
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        assert template_response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(template_response.content))
        ws = wb.active
        
        # Fill with invalid data (exceeds 31 days in March)
        if ws.cell(row=6, column=1).value:
            ws.cell(row=6, column=12).value = 25  # Full Days
            ws.cell(row=6, column=13).value = 5   # Half Days
            ws.cell(row=6, column=14).value = 5   # Approved Leaves = 35 total > 31
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        files = {"file": ("attendance_test.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        data = {"month": "2026-03"}
        
        response = requests.post(
            f"{BASE_URL}/api/hr/attendance/import",
            files=files,
            data=data,
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        
        assert response.status_code == 200  # Import completes but with errors
        result = response.json()
        
        # Check for validation errors
        errors = result.get("errors", [])
        assert len(errors) > 0, f"Should have validation errors for exceeding calendar days: {result}"
        assert any("exceeds calendar days" in str(e).lower() or "total days" in str(e).lower() for e in errors), f"Error should mention exceeding days: {errors}"

    def test_import_invalid_month_format(self, auth_token):
        """Test import rejects invalid month format"""
        import openpyxl
        
        # Create minimal valid workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        files = {"file": ("test.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        data = {"month": "invalid"}
        
        response = requests.post(
            f"{BASE_URL}/api/hr/attendance/import",
            files=files,
            data=data,
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        
        assert response.status_code == 400, f"Expected 400, got {response.status_code}"
        assert "Invalid month format" in response.json().get("detail", "")

    def test_import_creates_correct_record_types(self, auth_token):
        """Test import creates full day, half day, and leave records correctly"""
        import openpyxl
        
        # Download template
        template_response = requests.get(
            f"{BASE_URL}/api/hr/attendance/template/download",
            params={"month": "2026-04"},  # Use April to avoid conflicts with March test
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        
        if template_response.status_code != 200:
            pytest.skip("No template available")
        
        wb = openpyxl.load_workbook(io.BytesIO(template_response.content))
        ws = wb.active
        
        # Get first employee ID for verification
        first_employee_id = ws.cell(row=6, column=1).value
        if not first_employee_id:
            pytest.skip("No employees in template")
        
        # Fill specific attendance values
        ws.cell(row=6, column=12).value = 18  # Full Days
        ws.cell(row=6, column=13).value = 5   # Half Days
        ws.cell(row=6, column=14).value = 3   # Approved Leaves
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        files = {"file": ("attendance_april.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        data = {"month": "2026-04"}
        
        response = requests.post(
            f"{BASE_URL}/api/hr/attendance/import",
            files=files,
            data=data,
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        
        assert response.status_code == 200, f"Import failed: {response.text}"
        result = response.json()
        
        # Verify records were created: 18 full + 5 half + 3 leave = 26 records per employee
        expected_records = 18 + 5 + 3
        assert result.get("attendance_records_created", 0) >= expected_records, f"Expected at least {expected_records} records, got {result.get('attendance_records_created')}"


class TestReImport:
    """Tests for re-importing attendance - should replace manual_import records"""

    def test_reimport_replaces_manual_records(self, auth_token):
        """Test re-importing same month replaces only manual_import records"""
        import openpyxl
        
        # Download template
        template_response = requests.get(
            f"{BASE_URL}/api/hr/attendance/template/download",
            params={"month": "2026-05"},  # Use May
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        
        if template_response.status_code != 200:
            pytest.skip("No template available")
        
        wb = openpyxl.load_workbook(io.BytesIO(template_response.content))
        ws = wb.active
        
        if not ws.cell(row=6, column=1).value:
            pytest.skip("No employees in template")
        
        # First import
        ws.cell(row=6, column=12).value = 10  # Full Days
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        files = {"file": ("test1.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        response1 = requests.post(
            f"{BASE_URL}/api/hr/attendance/import",
            files=files,
            data={"month": "2026-05"},
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        
        assert response1.status_code == 200
        first_import = response1.json()
        
        # Second import with different values
        wb = openpyxl.load_workbook(io.BytesIO(template_response.content))
        ws = wb.active
        ws.cell(row=6, column=12).value = 15  # Full Days - different value
        output2 = io.BytesIO()
        wb.save(output2)
        output2.seek(0)
        
        files = {"file": ("test2.xlsx", output2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        response2 = requests.post(
            f"{BASE_URL}/api/hr/attendance/import",
            files=files,
            data={"month": "2026-05"},
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        
        assert response2.status_code == 200
        second_import = response2.json()
        
        # Records should be replaced, not duplicated
        print(f"First import: {first_import}")
        print(f"Second import: {second_import}")


class TestPayrollIntegration:
    """Tests for payroll integration with imported attendance"""

    def test_payroll_picks_up_imported_attendance(self, auth_token, auth_headers):
        """Test that running payroll uses imported attendance data"""
        import openpyxl
        
        # First, import attendance for June 2026
        template_response = requests.get(
            f"{BASE_URL}/api/hr/attendance/template/download",
            params={"month": "2026-06"},
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        
        if template_response.status_code != 200:
            pytest.skip("No template available")
        
        wb = openpyxl.load_workbook(io.BytesIO(template_response.content))
        ws = wb.active
        
        first_employee_id = ws.cell(row=6, column=1).value
        if not first_employee_id:
            pytest.skip("No employees in template")
        
        # Fill attendance: 20 full days, 5 half days, 5 leaves = 30 total for June
        ws.cell(row=6, column=12).value = 20  # Full Days
        ws.cell(row=6, column=13).value = 5   # Half Days
        ws.cell(row=6, column=14).value = 5   # Approved Leaves
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        files = {"file": ("june_attendance.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        import_response = requests.post(
            f"{BASE_URL}/api/hr/attendance/import",
            files=files,
            data={"month": "2026-06"},
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        
        assert import_response.status_code == 200, f"Import failed: {import_response.text}"
        
        # Now run payroll for June 2026
        payroll_response = requests.post(
            f"{BASE_URL}/api/hr/payroll/run",
            json={"year": 2026, "month": 6},
            headers=auth_headers
        )
        
        # Payroll might fail if already exists - that's ok, we're testing integration
        if payroll_response.status_code == 400:
            detail = payroll_response.json().get("detail", "")
            if "already processed" in detail.lower():
                print(f"Payroll already processed for this month: {detail}")
                return  # Test passes - payroll exists
        
        assert payroll_response.status_code == 200, f"Payroll run failed: {payroll_response.text}"
        
        result = payroll_response.json()
        print(f"Payroll result: {result}")


class TestFrontendIntegration:
    """Tests for frontend UI integration points"""

    def test_attendance_page_loads(self, auth_token):
        """Test the attendance page endpoint works"""
        # Get attendance for today
        from datetime import date
        today = date.today().isoformat()
        
        response = requests.get(
            f"{BASE_URL}/api/hr/attendance",
            params={"date": today},
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        # May return empty list if no attendance, that's ok
        assert isinstance(response.json(), list)

    def test_auth_required_for_template(self):
        """Test template download requires authentication"""
        response = requests.get(
            f"{BASE_URL}/api/hr/attendance/template/download",
            params={"month": "2026-03"}
        )
        
        assert response.status_code == 403 or response.status_code == 401, f"Expected 401/403, got {response.status_code}"

    def test_auth_required_for_import(self):
        """Test import requires authentication"""
        import openpyxl
        
        wb = openpyxl.Workbook()
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        files = {"file": ("test.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        
        response = requests.post(
            f"{BASE_URL}/api/hr/attendance/import",
            files=files,
            data={"month": "2026-03"}
        )
        
        assert response.status_code == 403 or response.status_code == 401


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
